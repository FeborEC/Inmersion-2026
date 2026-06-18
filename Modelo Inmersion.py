#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
============================================================================
 FEBOR · Inmersión 2026 — Generador del buscador documental y de actas
============================================================================

Este script NO reescribe tu aplicativo. Tu 'index.html' es un
bundle autocontenido (diseño + PDFs embebidos en base64). Lo que hace este
script es REGENERAR los bloques de datos que tu HTML ya consume y volverlos
a inyectar en su sitio, conservando intacto todo tu diseño.

Bloques que regenera e inyecta dentro del HTML:
  · docIdxB64   -> índice de búsqueda  {docs:[{id,title,text,top}], global:[...]}
  · actasB64    -> lista de actas      [{year,month,acta,fecha,sesion,temas}]
  · <id>B64     -> el PDF de cada documento embebido en base64
                   (cbcfB64, cbcf2B64..cbcf5B64, estatutoB64, codigoB64, eticaB64)

Flujo (un solo comando):
    python generar_indice.py

Pasos:
  1. Lee los PDF de las cuatro carpetas y los asocia a los IDs del aplicativo.
  2. Extrae texto, calcula el índice (frecuencias) -> docIdxB64
  3. Embebe cada PDF en base64 -> <id>B64
  4. Lee 'Relación de Actas Consejo.xlsx' (solo la RELACIÓN) -> actasB64
  5. Inyecta todo dentro de 'index.html' (queda listo para publicar).
  6. Publica en GitHub (git add / commit / push).

Requisitos:  pip install pypdf openpyxl

NOTA: no usa IA. Solo extrae texto, cuenta palabras y arma el índice;
es el mismo cálculo que hacía el aplicativo, pero en Python para que puedas
regenerarlo al cambiar PDFs o actas.
============================================================================
"""

import sys
import re
import json
import base64
import unicodedata
import subprocess
from pathlib import Path
from collections import Counter
from datetime import datetime

# ---------------------------------------------------------------------------
# Dependencias
# ---------------------------------------------------------------------------
try:
    from pypdf import PdfReader
except ImportError:
    sys.exit("Falta pypdf:  pip install pypdf")

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl:  pip install openpyxl")

# ---------------------------------------------------------------------------
# Rutas (TODO relativo a la raíz del proyecto, con pathlib)
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent

CARPETA_LEYES   = "1. Leyes_Decretos_Otros"
CARPETA_CBCF    = "2. C_Basica_Contable_Financiera"
CARPETA_JURID   = "3. C_Basica_Juridica"
CARPETA_INTERN  = "4. Normas_Internas"

CARPETAS = [CARPETA_LEYES, CARPETA_CBCF, CARPETA_JURID, CARPETA_INTERN]

EXCEL_ACTAS = "Relación de Actas Consejo.xlsx"
HTML_FILE   = "index.html"

# ---------------------------------------------------------------------------
#  CATÁLOGO DE DOCUMENTOS
#  Mapea cada ID interno del aplicativo a: título, variable de blob y archivo.
#
#  - id        : el identificador que usa tu HTML (t1..t5, i1..i3, ...)
#  - title     : el título que se muestra en el índice de búsqueda
#  - blobvar   : la variable <X>B64 del HTML donde va embebido el PDF
#  - carpeta   : carpeta donde está el PDF
#  - patrones  : posibles nombres de archivo (se toma el primero que exista;
#                no distingue mayúsculas/minúsculas, ni guiones/espacios)
#
#  Si agregas documentos nuevos al aplicativo, añádelos aquí con su id y
#  su blobvar correspondiente.
# ---------------------------------------------------------------------------
CATALOGO = [
    # --- Circular Básica Contable y Financiera (carpeta 2) ---
    {"id": "t1", "title": "Título I: Disposiciones Comunes",
     "blobvar": "cbcfB64",  "carpeta": CARPETA_CBCF,
     "patrones": ["cbcf-t1", "titulo i", "titulo 1", "disposiciones comunes"]},
    {"id": "t2", "title": "Título II: Presentación de Estados Financieros",
     "blobvar": "cbcf2B64", "carpeta": CARPETA_CBCF,
     "patrones": ["cbcf-t2", "titulo ii", "titulo 2", "estados financieros"]},
    {"id": "t3", "title": "Título III: Régimen Prudencial",
     "blobvar": "cbcf3B64", "carpeta": CARPETA_CBCF,
     "patrones": ["cbcf-t3", "titulo iii", "titulo 3", "regimen prudencial"]},
    {"id": "t4", "title": "Título IV: Sistema de Administración de Riesgos",
     "blobvar": "cbcf4B64", "carpeta": CARPETA_CBCF,
     "patrones": ["cbcf-t4", "titulo iv", "titulo 4", "administracion de riesgos"]},
    {"id": "t5", "title": "Título V: Indicadores Financieros",
     "blobvar": "cbcf5B64", "carpeta": CARPETA_CBCF,
     "patrones": ["cbcf-t5", "titulo v", "titulo 5", "indicadores financieros"]},
    # --- Normas Internas (carpeta 4) ---
    {"id": "i1", "title": "Estatuto Social",
     "blobvar": "estatutoB64", "carpeta": CARPETA_INTERN,
     "patrones": ["estatuto"]},
    {"id": "i2", "title": "Código de Buen Gobierno",
     "blobvar": "codigoB64",  "carpeta": CARPETA_INTERN,
     "patrones": ["buen gobierno", "codigo de buen gobierno"]},
    {"id": "i3", "title": "Código de Ética y Buena Conducta",
     "blobvar": "eticaB64",   "carpeta": CARPETA_INTERN,
     "patrones": ["etica", "buena conducta", "codigo de etica"]},
]

# ---------------------------------------------------------------------------
# Texto / stopwords (mismo criterio del aplicativo)
# ---------------------------------------------------------------------------
STOP = set((
    "para con los las del una unos unas por como mas pero sus est este esta estos estas que "
    "ser son fue han hay sobre entre cada segun cuando donde quien cual sino antes desde hasta "
    "tras muy tambien esto eso aquel cuyo cuya debe deben podra podran dicha dicho dichas dichos "
    "mediante traves numeral literal articulo articulos paragrafo capitulo titulo demas cualquier "
    "todo toda todos todas otro otra otros otras dentro tal tales ello ellas ellos esa ese asi "
    "aqui cuales cuyos correspondiente correspondientes respectivo siguiente siguientes mismo "
    "misma debera deberan mismas mismos circular basica contable financiera organizacion "
    "organizaciones solidaria solidarias superintendencia solidario presente"
).split())

MESES_ABBR = {
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "set": 9, "oct": 10, "nov": 11, "dic": 12,
}
MESES_FULL = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10,
    "noviembre": 11, "diciembre": 12,
}


def fold(s: str) -> str:
    """minúsculas sin tildes."""
    return "".join(
        c for c in unicodedata.normalize("NFD", str(s).lower())
        if unicodedata.category(c) != "Mn"
    )


# ---------------------------------------------------------------------------
# Listar PDFs de una carpeta SIN duplicados por mayúsculas/minúsculas.
# En Windows el sistema de archivos no distingue mayúsculas, así que
# glob("*.pdf") + glob("*.PDF") devolvía el mismo archivo dos veces. Aquí se
# deduplica por nombre normalizado (en minúscula) conservando una sola entrada.
# ---------------------------------------------------------------------------
def listar_pdfs(dir_path):
    vistos = {}
    for pdf in list(dir_path.glob("*.pdf")) + list(dir_path.glob("*.PDF")):
        clave = pdf.name.lower()
        if clave not in vistos:
            vistos[clave] = pdf
    return list(vistos.values())


# ---------------------------------------------------------------------------
# Localizar el PDF de cada documento del catálogo
# ---------------------------------------------------------------------------
def encontrar_pdf(carpeta: str, patrones):
    dir_path = ROOT / carpeta
    if not dir_path.exists():
        return None
    pdfs = listar_pdfs(dir_path)
    for pat in patrones:
        pat_f = fold(pat).replace("-", " ").replace("_", " ")
        for pdf in pdfs:
            nombre = fold(pdf.stem).replace("-", " ").replace("_", " ")
            if pat_f in nombre:
                return pdf
    return None


# ---------------------------------------------------------------------------
# Formato de título jurídico:
#   'TITULO I ECONOMIA SOLIDARIA Y ORGANIZACIONES'  (incluso 'TUTULO VII ...')
#   ->  'Título I: Economia solidaria y organizaciones'
#   (numeral romano en mayúscula, tema en minúscula con inicial mayúscula)
# ---------------------------------------------------------------------------
def formato_titulo_juridico(stem):
    s = stem.strip()
    m = re.match(r"^\s*(t[iu]tulo)\s+([ivxlcdm]+)\s+(.*)$", s, re.IGNORECASE)
    if not m:
        return s[:1].upper() + s[1:].lower()
    romano = m.group(2).upper()
    tema = m.group(3).strip()
    tema_fmt = (tema[:1].upper() + tema[1:].lower()) if tema else ""
    return f"Título {romano}: {tema_fmt}"


def _romano_a_int(r):
    valores = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    total, prev = 0, 0
    for ch in reversed(r.upper()):
        v = valores.get(ch, 0)
        total += -v if v < prev else v
        prev = max(prev, v)
    return total


# ---------------------------------------------------------------------------
# 1-3. Índice de búsqueda + PDFs embebidos
# ---------------------------------------------------------------------------
def construir_indice_y_blobs():
    docs, gfreq, gdocs = [], Counter(), {}
    blobs = {}   # blobvar -> base64 del PDF

    for item in CATALOGO:
        pdf = encontrar_pdf(item["carpeta"], item["patrones"])
        if not pdf:
            print(f"  [aviso] sin PDF para {item['id']} ({item['title']}) "
                  f"en '{item['carpeta']}'. Se conserva el embebido anterior si existe.")
            continue

        try:
            reader = PdfReader(str(pdf))
            text = " ".join((pg.extract_text() or "") for pg in reader.pages)
        except Exception as e:
            print(f"  [error] no se pudo leer texto de {pdf.name}: {e}")
            text = ""
        text = re.sub(r"\s+", " ", text).strip()

        freq = Counter(
            w for w in re.findall(r"[a-zñ]{5,}", fold(text)) if w not in STOP
        )
        top = [{"w": w, "n": n} for w, n in freq.most_common(14)]
        for w, n in freq.items():
            if n >= 4:
                gfreq[w] += n
                gdocs.setdefault(w, set()).add(item["id"])

        docs.append({
            "id": item["id"],
            "title": item["title"],
            "text": text,
            "top": top,
        })

        blobs[item["blobvar"]] = base64.b64encode(pdf.read_bytes()).decode("ascii")
        print(f"  [ok] {item['id']:>3}  {item['title'][:42]:<42} "
              f"({len(text):>7} chars · {pdf.stat().st_size // 1024} KB)")

    glob = [
        {"w": w, "n": n, "docs": sorted(gdocs[w])}
        for w, n in gfreq.most_common(44)
    ]
    indice = {"docs": docs, "global": glob}
    print(f"  -> {len(docs)} documentos indexados · {len(blobs)} PDFs embebidos")
    return indice, blobs


# ---------------------------------------------------------------------------
# SECCIONES DINÁMICAS DE RECURSOS
# Cada carpeta documental se asocia a una sección del buscador. El script
# descubre TODOS los PDF de la carpeta, los embebe, indexa su texto y genera
# las tarjetas. Para agregar un documento: copiar el PDF a la carpeta y correr.
#
# Campos de cada sección:
#   - grupo    : título EXACTO del grupo en recGroups (para localizarlo en HTML)
#   - carpeta  : carpeta documental
#   - prefijo  : prefijo de id para el índice (debe ser único por sección)
#   - mapavar  : nombre de la variable JS donde van los PDF de esa sección
#   - color    : color del badge PDF
#   - meta     : texto bajo el nombre ('PDF · Norma externa', etc.)
#   - formato  : 'juridico' (Título X: tema) o 'archivo' (nombre con inicial mayús.)
#   - nuevo    : True si la sección NO existe aún y hay que crearla en el HTML
#   - antes_de : (solo si nuevo) título del grupo ANTES del cual insertar
# ---------------------------------------------------------------------------
SECCIONES = [
    {"grupo": "Leyes, decretos y otros", "carpeta": CARPETA_LEYES,
     "prefijo": "l", "mapavar": "leyesFilesB64", "color": "#7a5cc7",
     "meta": "PDF \u00b7 Norma externa", "formato": "archivo",
     "nuevo": True, "antes_de": "Circular Básica Contable y Financiera"},
    {"grupo": "Circular Básica Contable y Financiera", "carpeta": CARPETA_CBCF,
     "prefijo": "c", "mapavar": "contableFilesB64", "color": "#2f7be0",
     "meta": "PDF \u00b7 Norma externa", "formato": "juridico", "nuevo": False},
    {"grupo": "Circular Básica Jurídica", "carpeta": CARPETA_JURID,
     "prefijo": "j", "mapavar": "juridicaFilesB64", "color": "#2f7be0",
     "meta": "PDF \u00b7 Norma externa", "formato": "juridico", "nuevo": False},
    {"grupo": "Normas internas", "carpeta": CARPETA_INTERN,
     "prefijo": "n", "mapavar": "internasFilesB64", "color": "#e2574c",
     "meta": "PDF \u00b7 Documento oficial", "formato": "archivo", "nuevo": False},
]


def _formato_archivo(stem):
    """Nombre del archivo con SOLO la inicial en mayúscula (resto minúscula)."""
    s = re.sub(r"[_]+", " ", stem).strip()
    return s[:1].upper() + s[1:].lower() if s else s


def construir_seccion(sec):
    """Descubre los PDF de una carpeta y devuelve:
    (docs_indice, files_b64, items, glob_extra)."""
    dir_path = ROOT / sec["carpeta"]
    docs, files_b64, items = [], {}, []
    gfreq, gdocs = Counter(), {}

    if not dir_path.exists():
        print(f"  [aviso] no existe la carpeta '{sec['carpeta']}'.")
        return docs, files_b64, items, []

    pdfs = sorted(listar_pdfs(dir_path))
    if not pdfs:
        print(f"  [info] sin PDFs en '{sec['carpeta']}'.")
        return docs, files_b64, items, []

    # Orden: por número romano del título cuando exista; si no, alfabético.
    def clave_orden(p):
        m = re.match(r"^\s*t[iu]tulo\s+([ivxlcdm]+)\s", p.stem, re.IGNORECASE)
        return (0, _romano_a_int(m.group(1))) if m else (1, fold(p.stem))

    pdfs.sort(key=clave_orden)

    for idx, pdf in enumerate(pdfs, 1):
        if sec["formato"] == "juridico":
            title = formato_titulo_juridico(pdf.stem)
        else:
            title = _formato_archivo(pdf.stem)
        doc_id = f"{sec['prefijo']}{idx}"

        try:
            reader = PdfReader(str(pdf))
            text = " ".join((pg.extract_text() or "") for pg in reader.pages)
        except Exception as e:
            print(f"  [error] no se pudo leer {pdf.name}: {e}")
            text = ""
        text = re.sub(r"\s+", " ", text).strip()

        freq = Counter(
            w for w in re.findall(r"[a-zñ]{5,}", fold(text)) if w not in STOP
        )
        top = [{"w": w, "n": n} for w, n in freq.most_common(14)]
        for w, n in freq.items():
            if n >= 4:
                gfreq[w] += n
                gdocs.setdefault(w, set()).add(doc_id)

        docs.append({"id": doc_id, "title": title, "text": text, "top": top})
        files_b64[pdf.name] = base64.b64encode(pdf.read_bytes()).decode("ascii")
        items.append({"name": title, "fname": pdf.name, "key": pdf.name})
        print(f"  [ok] {sec['grupo'][:28]:<28} {idx}: {title[:42]}")

    glob_extra = [
        {"w": w, "n": n, "docs": sorted(gdocs[w])}
        for w, n in gfreq.most_common(44)
    ]
    print(f"  -> {len(items)} documentos en '{sec['grupo']}'")
    return docs, files_b64, items, glob_extra



# ---------------------------------------------------------------------------
def _norm_header(s):
    return fold(str(s or "").strip())


def _mes_desde_texto(s):
    f = fold(s)
    for k, v in MESES_FULL.items():
        if k in f:
            return v
    for k, v in MESES_ABBR.items():
        if re.search(r"\b" + k, f):
            return v
    return None


def _parse_fecha(valor):
    """Devuelve (texto_fecha, year_str, month_int|None) respetando el
    formato del aplicativo ('28 abr 2023')."""
    if valor is None or valor == "":
        return ("", "", None)

    if isinstance(valor, datetime):
        meses = ["", "ene", "feb", "mar", "abr", "may", "jun",
                 "jul", "ago", "sep", "oct", "nov", "dic"]
        txt = f"{valor.day:02d} {meses[valor.month]} {valor.year}"
        return (txt, str(valor.year), valor.month)

    s = str(valor).strip()
    m = re.search(r"(20\d{2}|19\d{2})", s)
    year = m.group(1) if m else ""
    month = _mes_desde_texto(s)
    if month is None:
        m2 = re.search(r"\b\d{1,2}[/-](\d{1,2})[/-]\d{2,4}\b", s)
        if m2:
            mm = int(m2.group(1))
            month = mm if 1 <= mm <= 12 else None
    return (s, year, month)


_PREFIJOS_NO_CORTAR = {"pre", "post", "ex", "anti", "sub", "vice", "co", "re"}


def _cap_inicial(s):
    """Pone en mayúscula la primera letra del texto, sin alterar el resto."""
    s = s.strip(" .-\t")
    if not s:
        return s
    for i, ch in enumerate(s):
        if ch.isalpha():
            return s[:i] + ch.upper() + s[i + 1:]
    return s


def _split_temas(valor):
    """Separa los temas de una acta en una lista (un tema por elemento),
    y deja cada tema con MAYÚSCULA inicial.

    En el Excel los temas vienen pegados con guion. Se separan así,
    PROTEGIENDO los guiones que son parte legítima del texto:
      · rangos de años/números:  '2022-2026', '2026 - 2029'  -> NO se parten
      · cédulas / numeraciones:  'CC. 73.105.747'             -> NO se parten
      · prefijos:                'Pre-Aprobación', 'PRE-REGISTRO' -> NO se parten

    Reglas de corte:
      A) '.-'  (punto-guion)                          -> siempre separa
      B) palabra/letra - palabra/letra                -> separa
         (p.ej. 'enero-Aprobacion', 'informe-revisión')
      C) número - MAYÚSCULA                           -> separa
         (p.ej. '2026- ELECCION'), pero número-número NO (rango)
      D) salto de línea, ';' y viñetas                -> separa
    """
    if valor is None:
        return []
    s = str(valor).strip()
    if not s:
        return []

    SEP = "\u0001"  # marcador temporal
    s = re.sub(r"\s*\n\s*", " ", s)        # unificar saltos de línea
    s = re.sub(r"\s*\.\-\s*", SEP, s)      # A) '.-' siempre separa

    # B y C) guion entre dos tokens (letras/números)
    def _corte(m):
        izq = m.group(1)         # palabra o número antes del guion
        der = m.group(2)         # primer carácter después del guion
        # rango numérico  dígito-dígito  -> NO cortar
        if izq[-1].isdigit() and der.isdigit():
            return m.group(0)
        # prefijo legítimo (Pre-, Post-, ...) -> NO cortar
        if izq.lower() in _PREFIJOS_NO_CORTAR:
            return m.group(0)
        # letra-letra  -> cortar
        if izq[-1].isalpha() and der.isalpha():
            return izq + SEP + der
        # número-MAYÚSCULA  -> cortar
        if izq[-1].isdigit() and der.isupper():
            return izq + SEP + der
        return m.group(0)

    s = re.sub(
        r"([A-Za-zÁÉÍÓÚáéíóúÑñ0-9]+)\s*\-\s*([A-Za-zÁÉÍÓÚáéíóúÑñ0-9])",
        _corte, s,
    )
    s = re.sub(r"\s*[;•·]\s*", SEP, s)     # D) ';' y viñetas

    # Unificar TODOS los temas en MAYÚSCULAS (para que no se vea desordenado:
    # en el Excel unos vienen en mayúscula y otros en minúscula).
    partes = [p.strip(" .-\t").upper() for p in s.split(SEP)]
    return [p for p in partes if p]


#
# El Excel tiene UNA HOJA POR AÑO, con formatos distintos:
#   · "AÑO 2021"  -> encabezado en otra fila; columnas: No. ACTA | TEMA
#                    (sin fecha ni sesión; un tema por fila, el número se repite)
#   · "AÑO 2022"  -> FECHA | No. ACTA | TEMA (algunas fechas vacías)
#   · "ACTAS 2023/2024/2025/2026" -> FECHA | No. ACTA | SESION | TEMA
# Se omite "RESOLUCIONES 2023" (son resoluciones, no actas).
#
# Hojas de actas a procesar (resto se omite automáticamente):
HOJAS_ACTAS = [
    "AÑO 2021", "AÑO 2022",
    "ACTAS 2023", "ACTAS 2024", "ACTAS 2025", "ACTAS 2026",
]


def _detectar_encabezado(rows):
    """Encuentra la fila de encabezado real: la que tiene la columna TEMA
    junto a la columna de número de acta (evita la fila de título)."""
    for i, r in enumerate(rows):
        celdas = [_norm_header(c) for c in r if c not in (None, "")]
        tiene_tema = any("tema" in c or "asunto" in c for c in celdas)
        tiene_acta = any(("acta" in c or c.strip() in ("no.", "no")) for c in celdas)
        if tiene_tema and tiene_acta:
            return i
    return None


def construir_actas():
    xlsx_path = ROOT / EXCEL_ACTAS
    if not xlsx_path.exists():
        print(f"  [aviso] no se encontró '{EXCEL_ACTAS}'. Se omiten actas.")
        return []

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    # Diccionario ordenado: clave (año, acta) -> registro agrupado.
    # Así, si un mismo número de acta aparece en varias filas (2021/2022),
    # se juntan todos sus temas en una sola acta.
    from collections import OrderedDict
    registros = OrderedDict()

    for sn in wb.sheetnames:
        if sn not in HOJAS_ACTAS:
            print(f"  [skip] hoja '{sn}' (no es de actas)")
            continue

        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True, max_col=8))
        if not rows:
            continue

        # Año tomado del nombre de la hoja (respaldo cuando la fecha viene vacía)
        ym = re.search(r"(20\d{2})", sn)
        anio_hoja = ym.group(1) if ym else ""

        hidx = _detectar_encabezado(rows)
        if hidx is None:
            print(f"  [!!] '{sn}': no se halló encabezado, se omite.")
            continue

        headers = [_norm_header(c) if c else "" for c in rows[hidx]]

        def col(*claves):
            for idx, h in enumerate(headers):
                for k in claves:
                    if k in h:
                        return idx
            return None

        c_acta   = col("acta", "numero", "num")
        c_fecha  = col("fecha")
        c_sesion = col("sesion", "tipo", "clase")
        c_temas  = col("tema", "asunto", "orden", "tratado")

        if c_acta is None:
            print(f"  [!!] '{sn}': sin columna de acta, se omite.")
            continue

        n_hoja = 0
        ultimo_acta = None  # para arrastrar el número en filas que solo traen tema

        for r in rows[hidx + 1:]:
            if all(c in (None, "") for c in r):
                continue

            def get(idx):
                return r[idx] if (idx is not None and idx < len(r)) else None

            acta_no = get(c_acta)
            acta_no = re.sub(r"\.0$", "", str(acta_no).strip()) if acta_no not in (None, "") else ""

            fecha_txt, year, month = _parse_fecha(get(c_fecha))
            if not year:
                year = anio_hoja

            sesion = get(c_sesion)
            sesion = str(sesion).strip() if sesion not in (None, "") else ""
            temas = _split_temas(get(c_temas))

            # Fila sin número de acta: pertenece a la acta anterior (continúa temas)
            if not acta_no:
                if not temas:
                    continue
                acta_no = ultimo_acta
            if not acta_no:
                continue
            ultimo_acta = acta_no

            clave = (year, acta_no)
            if clave not in registros:
                registros[clave] = {
                    "year": year,
                    "month": month,       # int o None (el HTML acepta null)
                    "acta": acta_no,
                    "fecha": fecha_txt,
                    "sesion": sesion,
                    "temas": [],
                }
            reg = registros[clave]
            if not reg["fecha"] and fecha_txt:
                reg["fecha"] = fecha_txt
            if reg["month"] is None and month:
                reg["month"] = month
            if not reg["sesion"] and sesion:
                reg["sesion"] = sesion
            for t in temas:
                if t not in reg["temas"]:
                    reg["temas"].append(t)
            n_hoja += 1

        print(f"  [ok] '{sn}': {n_hoja} filas (año {anio_hoja})")

    actas = list(registros.values())

    def keyf(a):
        try:
            n = int(re.sub(r"\D", "", a["acta"]) or 0)
        except ValueError:
            n = 0
        return (a["year"] or "0", n)
    actas.sort(key=keyf)

    print(f"  -> {len(actas)} actas en total (agrupadas por número)")
    return actas


# ---------------------------------------------------------------------------
# 5. Inyectar los bloques de datos dentro del HTML existente
# ---------------------------------------------------------------------------
def _b64(obj):
    return base64.b64encode(
        json.dumps(obj, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")


def _reemplazar_asignacion(html, varname, nuevo_b64):
    """Reemplaza  <varname> = '....'  por el nuevo valor base64,
    conservando el resto del HTML. Devuelve (html, ok)."""
    patron = re.compile(r"(" + re.escape(varname) + r"\s*=\s*')[A-Za-z0-9+/=]*(')")
    if not patron.search(html):
        return html, False
    nuevo = patron.sub(lambda m: m.group(1) + nuevo_b64 + m.group(2), html, count=1)
    return nuevo, True


def _compactar_timeline(html):
    """Reorganiza la línea de tiempo de actas para aprovechar el ancho:
    las tarjetas pasan de flex-wrap (3-4 por fila pegadas a la izquierda)
    a un grid que llena todo el ancho disponible, con tarjetas más compactas
    y menos espacio vertical entre meses. Es idempotente (se puede correr
    varias veces sin acumular cambios)."""

    PATCHES = [
        # Contenedor de tarjetas POR MES: flex-wrap -> grid auto-fill
        ('display:flex; flex-wrap:wrap; gap:8px;\\">',
         'display:grid; grid-template-columns:repeat(auto-fill,minmax(62px,1fr)); gap:6px;\\">'),
        # Contenedor sin meses (años 2021/2022): flex-wrap -> grid auto-fill
        ('display:flex; flex-wrap:wrap; gap:8px; padding:14px 0 4px 12px;\\">',
         'display:grid; grid-template-columns:repeat(auto-fill,minmax(62px,1fr)); gap:6px; padding:10px 0 4px 12px;\\">'),
        # Tarjeta de acta -> más compacta (2 ocurrencias idénticas)
        ('gap:2px; min-width:58px; padding:9px 12px; border-radius:12px;',
         'gap:1px; min-width:0; padding:6px 5px; border-radius:9px;'),
        # Separación entre meses -> menor
        ('padding:14px 0 4px 20px; display:flex; flex-direction:column; gap:16px;',
         'padding:9px 0 4px 18px; display:flex; flex-direction:column; gap:9px;'),
        # Etiqueta de mes -> menos margen y tamaño
        ('font-size:12.5px; font-weight:800; color:{{ g.yColor }}; margin-bottom:9px;',
         'font-size:11.5px; font-weight:800; color:{{ g.yColor }}; margin-bottom:5px;'),
        # Subtexto 'tema(s)' -> un poco más chico (2 ocurrencias idénticas)
        ('font-size:9.5px; opacity:.7;',
         'font-size:8.5px; opacity:.65;'),
    ]

    n = 0
    for a, b in PATCHES:
        if a in html:
            html = html.replace(a, b)
            n += 1

    # Número del acta 14px -> 13px (solo el span que precede a {{ a.acta }})
    pat_num = re.compile(
        r'font-size:14px; font-weight:800;(?=[\\">]*\{\{ a\.acta \}\})'
    )
    if pat_num.search(html):
        html = pat_num.sub('font-size:13px; font-weight:800;', html)
        n += 1

    if n:
        print(f"  [ok] línea de tiempo compactada ({n} ajustes de estilo)")
    else:
        print("  [info] línea de tiempo: ya estaba compactada o no se halló el bloque")
    return html


def _compactar_cronograma(html):
    """Hace el calendario (Cronograma) más compacto: las celdas de día pasan
    de cuadradas (aspect-ratio 1) a más planas, y se reduce la separación.
    Idempotente."""
    PATCHES = [
        # Celda de día: cuadrada -> más plana
        ("aspectRatio: '1', borderRadius: '10px', display: 'flex', flexDirection: 'column', alignItems: 'center', justifyContent: 'center', fontSize: '12.5px'",
         "aspectRatio: '1.7', borderRadius: '10px', display: 'flex', flexDirection: 'column', alignItems: 'center', justifyContent: 'center', fontSize: '12px'"),
        # Celdas en blanco (relleno) también más planas
        ("if (cfg.blank) return React.createElement('div', { style: { aspectRatio: '1' } });",
         "if (cfg.blank) return React.createElement('div', { style: { aspectRatio: '1.7' } });"),
        # Grid de celdas: menos separación
        ('display:grid; grid-template-columns:repeat(7,1fr); gap:6px;\\">',
         'display:grid; grid-template-columns:repeat(7,1fr); gap:4px;\\">'),
        # Grid de encabezados (D L M ...): menos separación
        ('display:grid; grid-template-columns:repeat(7,1fr); gap:6px; margin-bottom:6px;\\">',
         'display:grid; grid-template-columns:repeat(7,1fr); gap:4px; margin-bottom:4px;\\">'),
    ]
    n = 0
    for a, b in PATCHES:
        if a in html:
            html = html.replace(a, b)
            n += 1
    if n:
        print(f"  [ok] cronograma compactado ({n} ajustes de estilo)")
    else:
        print("  [info] cronograma: ya estaba compactado o no se halló el bloque")
    return html


def _silenciar_errores_recursos(html):
    """Evita el banner rojo '[bundle] error'. El bundle trae un capturador
    global de errores que muestra CUALQUIER evento 'error' de la ventana,
    incluidos los inofensivos por carga de recursos externos (el iframe del
    tour 360°, imágenes, fuentes). Se filtra para que solo muestre errores
    reales de JavaScript (los que tienen mensaje). Idempotente."""
    OLD = (
        "window.addEventListener('error', function(e) {\n"
        "    var p = document.body || document.documentElement;\n"
        "    var d = document.getElementById('__bundler_err') || p.appendChild(document.createElement('div'));\n"
        "    d.id = '__bundler_err';"
    )
    NEW = (
        "window.addEventListener('error', function(e) {\n"
        "    var tgt = e && e.target;\n"
        "    if (tgt && tgt !== window && (tgt.tagName === 'IMG' || tgt.tagName === 'IFRAME' || tgt.tagName === 'SCRIPT' || tgt.tagName === 'LINK' || tgt.src || tgt.href)) return;\n"
        "    if (!e.message) return;\n"
        "    var p = document.body || document.documentElement;\n"
        "    var d = document.getElementById('__bundler_err') || p.appendChild(document.createElement('div'));\n"
        "    d.id = '__bundler_err';"
    )
    if OLD in html:
        html = html.replace(OLD, NEW)
        print("  [ok] banner '[bundle] error' silenciado (errores de recursos ignorados)")
    else:
        print("  [info] capturador de errores: ya estaba ajustado o no se halló")
    return html


def _mejorar_movil(html):
    """Mejoras para la vista en celular:
    1) Agrega el meta viewport (esencial: sin él, el móvil muestra todo
       diminuto a ancho de escritorio).
    2) Reduce el padding lateral de los contenedores principales en pantallas
       pequeñas (32px aprieta demasiado en un teléfono).
    Idempotente."""
    cambios = 0

    # 1) meta viewport tras el charset (en el <head> externo, no en el bundle)
    charset = '<meta charset="utf-8">'
    viewport = ('<meta charset="utf-8">\n  <meta name="viewport" '
                'content="width=device-width, initial-scale=1">')
    if 'name="viewport"' not in html and charset in html:
        html = html.replace(charset, viewport, 1)
        cambios += 1

    # 2) regla de padding para móvil dentro de la media query existente (600px)
    #    El bloque vive dentro de la cadena JSON del bundle -> usar \\n escapado.
    ancla_media = ('@media (max-width: 600px) {\\n  [data-mod-grid] '
                   '{ grid-template-columns: 1fr !important; }\\n  '
                   '[data-lower-grid] { grid-template-columns: 1fr !important; }')
    if ancla_media in html and "pad-movil-aplicado" not in html:
        regla = (ancla_media + '\\n  /* pad-movil-aplicado */\\n'
                 '  [style*=\\"max-width:1000px\\"] { padding-left:16px '
                 '!important; padding-right:16px !important; }')
        html = html.replace(ancla_media, regla, 1)
        cambios += 1

    if cambios:
        print(f"  [ok] mejoras de vista móvil aplicadas ({cambios})")
    else:
        print("  [info] vista móvil: ya estaba ajustada")
    return html


def _acordeon_resultados(html):
    """Convierte los resultados del asistente documental en un acordeón:
    cada documento aparece colapsado (solo título + N° coincidencias) y al
    hacer clic en la cabecera se despliegan los textos encontrados. Toca el
    JS (agrega 'expanded' y 'toggle' por documento) y el template (cabecera
    clickeable + chevron + envuelve snippets y botón en sc-if). Idempotente."""

    if "r.toggle" in html:
        print("  [info] acordeón de resultados: ya estaba aplicado")
        return html

    cambios = 0

    # --- 1) JS: agregar expanded + toggle a cada entrada de docResults ---
    push_anchor = ("open: () => this.openDocTab({ name: doc.title.replace(/:/g, "
                   "' -')")
    push_nuevo = (
        "expanded: !!(this.state.expandedDocs && this.state.expandedDocs[doc.id]), "
        "notExpanded: !(this.state.expandedDocs && this.state.expandedDocs[doc.id]), "
        "toggle: () => { const m = Object.assign({}, this.state.expandedDocs); "
        "m[doc.id] = !m[doc.id]; this.setState({ expandedDocs: m }); }, "
        "open: () => this.openDocTab({ name: doc.title.replace(/:/g, ' -')"
    )
    if push_anchor in html:
        html = html.replace(push_anchor, push_nuevo, 1)
        cambios += 1

    # --- 2) Template: cabecera clickeable + chevron ---
    header_open = ('<div style=\\"display:flex; align-items:center; gap:12px; '
                   'margin-bottom:9px;\\">')
    header_nuevo = ('<div onclick=\\"{{ r.toggle }}\\" style=\\"display:flex; '
                    'align-items:center; gap:12px; margin-bottom:9px; '
                    'cursor:pointer;\\">')
    if header_open in html:
        html = html.replace(header_open, header_nuevo, 1)
        cambios += 1

    # chevron tras el countLabel (icono según r.expanded / r.notExpanded)
    countlabel_close = '{{ r.countLabel }}<\\u002Fdiv>'
    chevron_cond = (
        '{{ r.countLabel }}<\\u002Fdiv>'
        '<sc-if value=\\"{{ r.expanded }}\\" hint-placeholder-val=\\"{{ false }}\\">'
        '<div style=\\"flex:0 0 auto; width:34px; height:34px; border-radius:50%; '
        'background:#eaf2fe; display:flex; align-items:center; '
        'justify-content:center; color:#2f7be0;\\">'
        '<svg width=\\"22\\" height=\\"22\\" viewBox=\\"0 0 24 24\\" fill=\\"none\\" '
        'stroke=\\"currentColor\\" stroke-width=\\"3.5\\" stroke-linecap=\\"round\\" '
        'stroke-linejoin=\\"round\\"><polyline points=\\"6 9 12 15 18 9\\">'
        '<\\u002Fpolyline><\\u002Fsvg><\\u002Fdiv><\\u002Fsc-if>'
        '<sc-if value=\\"{{ r.notExpanded }}\\" hint-placeholder-val=\\"{{ true }}\\">'
        '<div style=\\"flex:0 0 auto; width:34px; height:34px; border-radius:50%; '
        'background:#eef2f7; display:flex; align-items:center; '
        'justify-content:center; color:#2f7be0;\\">'
        '<svg width=\\"22\\" height=\\"22\\" viewBox=\\"0 0 24 24\\" fill=\\"none\\" '
        'stroke=\\"currentColor\\" stroke-width=\\"3.5\\" stroke-linecap=\\"round\\" '
        'stroke-linejoin=\\"round\\"><polyline points=\\"9 18 15 12 9 6\\">'
        '<\\u002Fpolyline><\\u002Fsvg><\\u002Fdiv><\\u002Fsc-if>'
    )
    if countlabel_close in html:
        html = html.replace(countlabel_close, chevron_cond, 1)
        cambios += 1

    # --- 3) Template: envolver snippets + botón en sc-if r.expanded ---
    snippets_open = ('<div data-scroll=\\"\\" style=\\"display:flex; '
                     'flex-direction:column; gap:8px; margin-bottom:12px;')
    snippets_nuevo = ('<sc-if value=\\"{{ r.expanded }}\\" '
                      'hint-placeholder-val=\\"{{ false }}\\">'
                      '<div data-scroll=\\"\\" style=\\"display:flex; '
                      'flex-direction:column; gap:8px; margin-bottom:12px;')
    if snippets_open in html:
        html = html.replace(snippets_open, snippets_nuevo, 1)
        cambios += 1

    abrir_btn_end = '{{ icoDown }} Abrir documento<\\u002Fdiv>'
    abrir_nuevo = '{{ icoDown }} Abrir documento<\\u002Fdiv><\\u002Fsc-if>'
    if abrir_btn_end in html:
        html = html.replace(abrir_btn_end, abrir_nuevo, 1)
        cambios += 1

    if cambios == 5:
        print("  [ok] acordeón de resultados aplicado (clic para desplegar)")
    else:
        print(f"  [!!] acordeón: solo {cambios}/5 cambios; revisar anclas")
    return html


def _quitar_scroll_doble(html):
    """Elimina la doble barra de desplazamiento del asistente de documentos.
    El panel del asistente (contenedor fijo) ya tiene su propio scroll; además,
    la lista interna de resultados tenía 'max-height + overflow-y:auto', lo que
    creaba una segunda barra pegada a la primera. Se le quita el scroll propio
    a la lista interna para que fluya y solo se desplace el panel. Idempotente."""
    target = ("display:flex; flex-direction:column; gap:8px; margin-bottom:12px; "
              "max-height:340px; overflow-y:auto; padding-right:6px; "
              "scrollbar-width:auto; scrollbar-color:#1bb59f transparent;")
    nuevo = ("display:flex; flex-direction:column; gap:8px; margin-bottom:12px; "
             "padding-right:6px;")
    if target in html:
        html = html.replace(target, nuevo, 1)
        print("  [ok] doble barra de desplazamiento corregida (asistente)")
    elif nuevo in html:
        print("  [info] doble barra: ya estaba corregida")
    else:
        print("  [info] doble barra: no se halló el contenedor interno")
    return html


def _quitar_botones_info(html):
    """Quita los botones circulares de información (ⓘ) que aparecían en las
    tarjetas de la Circular Contable. Esos botones solo se renderizan cuando
    a un item se le asigna '.info', cosa que hace un bucle con infoMap. Al
    dejar el infoMap vacío, ningún item recibe '.info' y el botón (que está
    bajo un sc-if value="{{ r.info }}") deja de aparecer. Idempotente."""
    target = "const infoMap = ['t1', 't2', 't3', 't4', 't5'];"
    if target in html:
        html = html.replace(target, "const infoMap = [];", 1)
        print("  [ok] botones de información (ⓘ) quitados de las tarjetas")
    elif "const infoMap = [];" in html:
        print("  [info] botones de información: ya estaban quitados")
    else:
        print("  [info] botones de información: no se halló el infoMap")
    return html


def _quitar_modulo3_velez(html):
    """Quita 'Módulo 3: Marco Legal' SOLO de la lista de temas de Juan Pablo
    Vélez. Se elimina la entrada con su coma final (que es exclusiva de su
    array, en medio de la lista); el docente dueño del Módulo 3 no se toca.
    Idempotente."""
    entrada = "'Módulo 3: Marco Legal: Reglas Claves para Actuar con Seguridad', "
    if entrada in html:
        html = html.replace(entrada, "", 1)
        print("  [ok] 'Módulo 3' quitado de los temas de Juan Pablo Vélez")
    else:
        print("  [info] 'Módulo 3' en Vélez: ya estaba quitado o no se halló")
    return html


def _esc_js(s):
    return s.replace("\\", "\\\\").replace("'", "\\'")


def _items_js_de_seccion(sec, items):
    """Construye el string JS del arreglo de items para una sección."""
    js = []
    for it in items:
        name = _esc_js(it["name"])
        fname = _esc_js(it["fname"])
        key = _esc_js(it["key"])
        js.append(
            "{ lead: this.fileBadge('PDF', '" + sec["color"] + "'), name: '" + name +
            "', meta: '" + sec["meta"] + "', tail: RI.down, action: () => "
            "this.openDocTab({ name: '" + name + "', fname: '" + fname +
            "', b64: (JSON.parse(atob(this." + sec["mapavar"] + "))['" + key + "']) }) }"
        )
    return ", ".join(js)


def _reemplazar_items_grupo(html, grupo_title, items_str):
    """Reemplaza el arreglo items: [...] de un grupo existente, contando
    corchetes balanceados (soporta corchetes anidados)."""
    marcador = grupo_title + "', icon: RI."
    pos = html.find(marcador)
    if pos < 0:
        return html, False
    pos_items = html.find("items: ", pos)
    if pos_items < 0:
        return html, False
    ini = html.find("[", pos_items)
    if ini < 0:
        return html, False
    depth, fin = 0, -1
    for k in range(ini, len(html)):
        if html[k] == "[":
            depth += 1
        elif html[k] == "]":
            depth -= 1
            if depth == 0:
                fin = k
                break
    if fin < 0:
        return html, False
    return html[:ini] + "[" + items_str + "]" + html[fin + 1:], True


def _inyectar_secciones(html, secciones_data):
    """Inyecta TODAS las secciones documentales dinámicas:
    - embebe el mapa base64 de cada sección como variable JS,
    - reemplaza/crea el arreglo items de cada grupo en recGroups.
    secciones_data: lista de (sec, items, files_b64).
    """
    SEP = "\\n  "  # separador ESCAPADO (texto barra+n), no salto real

    # ancla estable para insertar variables: la última var b64 conocida.
    m_anchor = re.search(r"(eticaB64 = '[A-Za-z0-9+/=]*';)", html)
    if not m_anchor:
        print("  [!!] no se encontró ancla para variables; se omite secciones.")
        return html
    ancla = m_anchor.group(1)

    bloque_vars = ""
    for sec, items, files_b64 in secciones_data:
        # 1) variable del mapa de archivos (idempotente: quitar previa)
        html = re.sub(
            r"(\\n  )?\s*" + re.escape(sec["mapavar"]) + r" = '[A-Za-z0-9+/=]*';",
            "", html,
        )
        mapa_b64 = base64.b64encode(
            json.dumps(files_b64, ensure_ascii=False).encode("utf-8")
        ).decode("ascii")
        bloque_vars += SEP + sec["mapavar"] + " = '" + mapa_b64 + "';"

    # insertar todas las variables juntas tras el ancla
    html = html.replace(ancla, ancla + bloque_vars, 1)

    # 2) items por grupo
    for sec, items, files_b64 in secciones_data:
        items_str = _items_js_de_seccion(sec, items)

        if sec.get("nuevo"):
            # crear el grupo nuevo ANTES del grupo indicado
            destino = sec["antes_de"]
            pos = html.find(destino + "', icon: RI.")
            if pos < 0:
                print(f"  [!!] no se pudo ubicar '{destino}' para crear "
                      f"'{sec['grupo']}'; se omite.")
                continue
            # retroceder hasta el '{' que abre ese grupo destino
            ini_grupo = html.rfind("{ title:", 0, pos)
            if ini_grupo < 0:
                print(f"  [!!] no se halló inicio del grupo destino; se omite "
                      f"'{sec['grupo']}'.")
                continue
            nuevo_grupo = (
                "{ title: '" + _esc_js(sec["grupo"]) +
                "', icon: RI.folder, items: [" + items_str + "] },\\n      "
            )
            # evitar duplicar si ya existe
            if (sec["grupo"] + "', icon: RI.") in html:
                html, ok = _reemplazar_items_grupo(html, sec["grupo"], items_str)
                print(f"  [ok] '{sec['grupo']}' actualizada ({len(items)} docs)")
            else:
                html = html[:ini_grupo] + nuevo_grupo + html[ini_grupo:]
                print(f"  [ok] '{sec['grupo']}' creada ({len(items)} docs)")
        else:
            html, ok = _reemplazar_items_grupo(html, sec["grupo"], items_str)
            estado = f"actualizada ({len(items)} docs)" if ok else "NO encontrada"
            print(f"  [{'ok' if ok else '!!'}] '{sec['grupo']}' {estado}")

    return html


def inyectar_en_html(indice, actas, blobs, secciones_data=None):
    html_path = ROOT / HTML_FILE
    if not html_path.exists():
        sys.exit(f"  [error] no se encontró '{HTML_FILE}' en la carpeta. "
                 "Coloca el HTML del aplicativo junto al script.")

    html = html_path.read_text(encoding="utf-8")

    cambios = []
    html, ok = _reemplazar_asignacion(html, "docIdxB64", _b64(indice))
    cambios.append(("docIdxB64", ok))
    html, ok = _reemplazar_asignacion(html, "actasB64", _b64(actas))
    cambios.append(("actasB64", ok))

    # Poblar TODAS las secciones documentales (dinámicas)
    html = _inyectar_secciones(html, secciones_data or [])

    # Reorganizar la línea de tiempo (diseño más compacto)
    html = _compactar_timeline(html)
    # Compactar el cronograma (calendario)
    html = _compactar_cronograma(html)
    # Quitar el banner rojo '[bundle] error' por recursos externos
    html = _silenciar_errores_recursos(html)
    # Quitar 'Módulo 3' de los temas de Juan Pablo Vélez
    html = _quitar_modulo3_velez(html)
    # Quitar los botones de información (ⓘ) de las tarjetas
    html = _quitar_botones_info(html)
    # Corregir la doble barra de desplazamiento del asistente
    html = _quitar_scroll_doble(html)
    # Acordeón: documentos colapsados, clic para desplegar
    html = _acordeon_resultados(html)
    # Mejoras para la vista en celular (viewport + padding)
    html = _mejorar_movil(html)

    # Escribir el HTML final sobre el mismo archivo (regenera en sitio).
    html_path.write_text(html, encoding="utf-8")

    for var, ok in cambios:
        estado = "actualizado" if ok else "NO encontrado en el HTML (se omitió)"
        print(f"  [{'ok' if ok else '!!'}] {var}: {estado}")
    print(f"  -> {HTML_FILE} generado en la carpeta del proyecto")


# ---------------------------------------------------------------------------
# 6. Publicación a GitHub
# ---------------------------------------------------------------------------
def publicar_github():
    print("\n== Publicando en GitHub ==")
    comandos = [
        ["git", "add", "."],
        ["git", "commit", "-m", "Actualizar buscador documental y actas"],
        ["git", "push"],
    ]
    for cmd in comandos:
        print("  $ " + " ".join(cmd))
        res = subprocess.run(cmd, cwd=str(ROOT), capture_output=True, text=True)
        salida = (res.stdout or "") + (res.stderr or "")
        if salida.strip():
            print("    " + salida.strip().replace("\n", "\n    "))
        if res.returncode != 0 and cmd[1] != "commit":
            print(f"    [aviso] '{cmd[1]}' terminó con código {res.returncode}.")


# ---------------------------------------------------------------------------
# Principal
# ---------------------------------------------------------------------------
def main():
    print(f"Raíz del proyecto: {ROOT}\n")

    print("== 1. Indexando documentos de TODAS las carpetas (dinámico) ==")
    # El índice de búsqueda se construye COMPLETO desde las secciones.
    indice = {"docs": [], "global": []}
    secciones_data = []

    for sec in SECCIONES:
        print(f"\n  -- {sec['grupo']}  ({sec['carpeta']}) --")
        docs, files_b64, items, glob_extra = construir_seccion(sec)
        indice["docs"].extend(docs)
        indice["global"].extend(glob_extra)
        secciones_data.append((sec, items, files_b64))

    # Recortar 'global' a los 44 términos más frecuentes (como el original)
    from collections import Counter as _C
    agg = _C()
    docs_by_word = {}
    for g in indice["global"]:
        agg[g["w"]] += g["n"]
        docs_by_word.setdefault(g["w"], set()).update(g["docs"])
    indice["global"] = [
        {"w": w, "n": n, "docs": sorted(docs_by_word[w])}
        for w, n in agg.most_common(44)
    ]

    print("\n== 4. Leyendo actas del Consejo ==")
    actas = construir_actas()

    print("\n== 5. Inyectando datos en el HTML ==")
    inyectar_en_html(indice, actas, {}, secciones_data)

    publicar_github()

    print("\n[LISTO] Abre el HTML localmente o revisa GitHub Pages.")


if __name__ == "__main__":
    main()
